from typing import List
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
import pytz
from datetime import datetime

from app_organization.models.organization import Organization
from app_system_setting.models import SystemApp, SystemMenu
from app_user.models.user import User, UserStatus

def CreateExcelTemplateSetting(title: str, sysAppStr: str, settingObjects):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = title
        # ใส่ header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        ws.cell(row=1, column=1).value = "Employee"
        ws.cell(row=1, column=1).alignment = styles.Alignment(horizontal='center')
        app: SystemApp = SystemApp.objects.filter(name = sysAppStr).first()
        header = ["code", "fNameEN", "lNameEN"]
        if app:
            menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
            if menus:
                ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=4+(len(menus)-1))
                ws.cell(row=1, column=4).value = "Menu"
                ws.cell(row=1, column=4).alignment = styles.Alignment(horizontal='center')
                # -- size menu column
                for col_idx in range(4, (4+len(menus))):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                for m in menus:
                    header.append(m.name)
        ws.append(header)
        # ใส่ข้อมูล
        user: list[User] = User.objects.filter(isActive = True, status = UserStatus.Hire.value)
        if user:
            dupUser = []
            if settingObjects:
                dupUser = [us.user.id for us in settingObjects]
            for u in user:
                if not u.id in dupUser:
                    ws.append([u.code, u.fNameEN, u.lNameEN])
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.row in [1,2]:
                    cell.alignment = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = styles.Font(bold=True)
                    cell.border = styles.Border(
                        left=styles.Side(style='thin'), 
                        right=styles.Side(style='thin'),
                        top=styles.Side(style='thin'), 
                        bottom=styles.Side(style='thin')
                    )
                else:
                    cell.alignment = styles.Alignment(wrap_text=True)
                    cell.border = styles.Border(
                        left=styles.Side(style='thin'), 
                        right=styles.Side(style='thin'),
                        top=styles.Side(style='thin'), 
                        bottom=styles.Side(style='thin')
                    )
        # -- size name column
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        return wb
    except Exception as e:
        print(e)
        return None
    
def printLogData(data):
    print("-------------------------------------------------------")
    print(f"print log : {data}")
    print("-------------------------------------------------------")
    
def DateStrToDate(date: str):
    tz = pytz.timezone("Asia/Bangkok")
    try:
        a = datetime.strptime(date, "%d/%m/%Y")
        b = tz.localize(a)
        return b.astimezone(pytz.utc)
    except Exception as e:
        print(e)
        return None
    
def dateStrAndTimeToDatetime(date: str, time: str):
    tz = pytz.timezone("Asia/Bangkok")
    try:
        a = datetime.strptime(date+" "+time, "%d/%m/%Y %H:%M")
        b = tz.localize(a)
        return b.astimezone(pytz.utc)
    except Exception as e:
        print(e)
        return None
    
def  getParentLevel(org: Organization, targetLevel: str):
    parent: Organization = org.parent
    while parent:
        # ตรวจสอบว่าระดับของ parent คือ Department หรือไม่
        if parent.level and parent.level.nameEN == targetLevel:
            return parent.serialize_organization()  # หรือ return parent.serialize_organization() ถ้าอยากได้ข้อมูลครบ
        parent = parent.parent
    return None  # ถ้าไม่มี parent ที่เป็น Department

def getChildLevel(org: Organization, targetLevel: str):
    results = []
    # หา child ทั้งหมดที่ parent = org
    children = Organization.objects.filter(isActive = True,parent=org)
    for child in children:
        if child.level and child.level.nameEN == targetLevel:
            # print(child.serialize_organization()['levelNameEN'])
            results.append(child.serialize_organization())
        # หาในระดับลึกต่อ (เช่น Factory → Department → Section)
        results.extend(getChildLevel(child, targetLevel))
    return results
    
def findDeptUser(cur: User):
    roles = []
    roleOthers = []
    roleDepts = []
    if not cur.roles:
        return roles
    
    for r in cur.roles:
        if r.isActive == True:
            if r.orgId.level.nameEN == "Department":
                roleDepts.append(r.orgId.serialize_organization())
            elif r.orgId.level.nameEN == "Managing Director":
                roleOthers.extend(getChildLevel(r.orgId, "Department"))
            elif r.orgId.level.nameEN == "Division":
                roleOthers.extend(getChildLevel(r.orgId, "Department"))
            else:
                roleOthers.append(getParentLevel(r.orgId, "Department"))

    if roleDepts:
        for d in roleDepts:
            if d not in roles:
                print('add dept')
                roles.append(d)

    if roleOthers:
        for r in roleOthers:
            if r not in roles:
                print('add other')
                roles.append(r)
    return roles