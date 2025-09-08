import tempfile
from typing import List
from bson import ObjectId
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from openpyxl import load_workbook
from app_organization.models.org_setting import OrgSetting
from app_organization.utils import HasOrgPermission, isSettingOrgAdmin
from app_system_setting.models import SystemApp, SystemMenu
from app_user.models.user import User
from app_user.utils import requiredLogin
from django.contrib import messages
from django.utils import timezone

from base_models.basemodel import UserSnapshot
from utilities.utility import CreateExcelTemplateSetting



@requiredLogin
def indexSettingOrg(request: HttpRequest):
    orgSettings = OrgSetting.objects.filter(isActive = True)
    hasPermission = HasOrgPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        print(f"has : {hasPermission}")
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    context ={
        "orgSettings": orgSettings
    }
    return render(request, 'setting_org/indexSetting.html', context)

@requiredLogin
def addSettingOrg(request: HttpRequest):
    response = HttpResponseRedirect(reverse('indexSettingOrg'))
    hasPermission = HasOrgPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    if request.method == "POST":
        try:
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            user = request.POST.get("user")
            if not user:
                messages.error(request, "User is required")
                return response
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            dupUser: OrgSetting = OrgSetting.objects.filter(user = user).first()
            if dupUser:
                messages.error(request, "Duplicate User")
                return response
            
            orgSetting = OrgSetting()
            orgSetting.user = ObjectId(user)
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            orgSetting.menus = menuList
            orgSetting.isAdmin = isadmin
            orgSetting.isActive = True
            currentUser: User = request.currentUser
            if currentUser:
                uCreate = UserSnapshot().UserToSnapshot(currentUser)
                if uCreate:
                    orgSetting.createBy = uCreate
            orgSetting.save()
            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response
    else:
        users: List[User] = User.objects.filter(isActive = True)
        app: SystemApp = SystemApp.objects.filter(name = "app_organization").first()
        if not app:
            messages.error(request, "App not found")
            return response
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        selectedUsers = []
        orgSettings: List[OrgSetting] = OrgSetting.objects.filter(isActive = True)
        if orgSettings:
            selectedUsers = [us.user.id for us in orgSettings]


        context = {
            "users": users,
            "menus": menus,
            "selectedUsers": selectedUsers
        }
        return render(request, 'setting_org/addSetting.html',context)
    
@requiredLogin
def editSettingOrg(request: HttpRequest, id: str):
    response = HttpResponseRedirect(reverse('indexSettingOrg'))
    hasPermission = HasOrgPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    if request.method == "POST":
        try:
            idsorg = request.POST.get("idsorg")
            if not idsorg:
                messages.error(request, "Id is required")
                return response
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            orgSetting: OrgSetting = OrgSetting.objects.filter(id = idsorg).first()
            if not orgSetting:
                messages.error(request, "Org Setting not found")
                return response
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            orgSetting.isAdmin = isadmin
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            orgSetting.menus = menuList
            orgSetting.updateDate = timezone.now()
            currentUser: User = request.currentUser
            if currentUser:
                uUpdateUser = UserSnapshot().UserToSnapshot(currentUser)
                if uUpdateUser:
                    orgSetting.updateBy = uUpdateUser
            orgSetting.save()

            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response
    else:
        if not id:
            messages.error(request, "Id is required")
            return response
        users: List[User] = User.objects.filter(isActive = True)
        app: SystemApp = SystemApp.objects.filter(name = "app_organization").first()
        if not app:
            messages.error(request, "App not found")
            return response
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        orgSetting: OrgSetting = OrgSetting.objects.filter(id = id).first()
        if not orgSetting:
            messages.error(request, "Org Setting not found")
            return response
        menu_ids = [m.id for m in orgSetting.menus]
        context = {
            "users": users,
            "menus": menus,
            "orgSetting": orgSetting,
            "menu_ids": menu_ids
        }
        return render(request, 'setting_org/editSetting.html', context)
    
@requiredLogin
def deleteSettingOrg(request: HttpRequest, id: str):
    try:
        hasPermission = HasOrgPermission(id = str(request.currentUser.id), checkAdmin = True)
        if not request.currentUser.isAdmin:
            if hasPermission == False:
                return JsonResponse({'deleted': False, 'message': 'Not Permission'})
        if not request.method == "GET":
            return JsonResponse({'deleted': False, 'message': 'Method not allowed'})
        if not id:
            return JsonResponse({'deleted': False, 'message': 'Not found id'})
        
        orgSetting: OrgSetting = OrgSetting.objects.filter(id = id).first()
        if not orgSetting:
            return JsonResponse({'deleted': False, 'message': 'Org Setting not found'})
        orgSetting.isActive = False
        orgSetting.updateDate = timezone.now()
        currentUser: User = request.currentUser
        if currentUser:
            uUpdate = UserSnapshot().UserToSnapshot(currentUser)
            if uUpdate:
                orgSetting.updateBy = uUpdate
        orgSetting.save()

        return JsonResponse({'deleted': True, 'message': 'Delete success'})
    except Exception as e:
        return JsonResponse({'deleted': False, 'message': str(e)})
    
@requiredLogin
def importSettingOrg(request: HttpRequest):
    hasPermission = HasOrgPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    if request.method == "POST":
        try:
            file = request.FILES.get("file")
            if not file:
                messages.error(request, "File is required")
                return HttpResponseRedirect(reverse('importSettingOrg'))
            if not file.name.endswith(".xlsx"):
                messages.error(request, "Only .xlsx files are supported.")
                return HttpResponseRedirect(reverse('importSettingOrg'))
            with tempfile.NamedTemporaryFile(delete=True, suffix=".xlsx") as temp_file:
                for chunk in file.chunks():
                    temp_file.write(chunk)
                # ✅ ข้อมูลถูกเขียนลงไฟล์ .xlsx บน disk จริง
                # flush() จะสั่งว่า: "เขียนเลยตอนนี้นะ!"
                temp_file.flush()
                # ย้าย “ตำแหน่ง pointer” ในไฟล์ — ว่าเราจะเริ่มอ่านหรือเขียนตรงไหน
                # หลังจากเขียนไฟล์เสร็จ pointer จะอยู่ "ท้ายไฟล์"
                # ถ้าจะเปิดอ่านต่อ (หรือใช้ file object เดิมอ่าน) ต้อง “ย้อนกลับไปที่ต้นไฟล์” → seek(0)
                temp_file.seek(0)
                wb = load_workbook(temp_file.name)
                ws = wb.active
                orgSettings = []
                header_row = [cell.value for cell in ws[2]]
                menu_names = header_row[3:]  # คอลัมน์เมนูเริ่มจาก index 4
                # ดึง SystemMenu ทั้งหมด
                system_menus = {menu.name: menu for menu in SystemMenu.objects(name__in=menu_names)}
                inserted_codes = set()
                rowCount = 3
                for row in ws.iter_rows(min_row=3):
                    code = str(row[0].value).strip() if row[0].value else None
                    if not code:
                        continue  # skip ถ้าไม่มี code 
                    if code in inserted_codes:
                        messages.error(request, f"Duplicate code: {code} in row ({rowCount})")# skip ถ้า code ซ้ำในไฟล์เดิม
                        return HttpResponseRedirect(reverse('importSettingOrg'))
                    user:User = User.objects.filter(code=code).first()
                    if not user:
                        messages.error(request, f"User not found in row ({rowCount}) with code: {code}")
                        return HttpResponseRedirect(reverse('importSettingOrg'))
                    checkOrgSetting: OrgSetting = OrgSetting.objects.filter(user = user).first()
                    if checkOrgSetting:
                        if checkOrgSetting.isActive == True:
                            # -- ถ้ายัง active ให้ show error
                            messages.error(request, f"Duplicate code: {code} in row ({rowCount})")
                            return HttpResponseRedirect(reverse('importSettingOrg'))
                        else:
                            # -- ถ้ามีข้อมูลใน db แล้ว ไม่ active false ให้ลบข้อมูลออกจาก db แล้ว save ของใหม่จาก excel ไปเลย
                            checkOrgSetting.delete()
                    selected_menu_names = []
                    for i, cell in enumerate(row[3:], start=3):
                        if str(cell.value).strip() == "1":
                            selected_menu_names.append(header_row[i])
                    if not selected_menu_names:
                        continue  # ข้ามถ้าไม่ได้เลือกเมนู
                    menu_refs = [system_menus[name] for name in selected_menu_names if name in system_menus]
                    # สร้าง OrgSetting
                    setting: OrgSetting = OrgSetting()
                    setting.user=user
                    setting.menus=menu_refs
                    setting.isActive=True
                    setting.isAdmin=False
                    setting.note=''
                    currentUser: User = request.currentUser
                    if currentUser:
                        uCreate = UserSnapshot().UserToSnapshot(currentUser)
                        if uCreate:
                            setting.createBy=uCreate
                    setting.createDate=timezone.now()
                    orgSettings.append(setting)
                    inserted_codes.add(code)
                    rowCount += 1
                # MongoEngine สามารถ insert หลาย document พร้อมกัน
                # ทำงานเร็วกว่า for setting in orgSettings: setting.save()
                OrgSetting.objects.insert(orgSettings)
            messages.success(request, 'Save Success')
            return HttpResponseRedirect(reverse('indexSettingOrg'))
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return HttpResponseRedirect(reverse('importSettingOrg'))
    else: 
        return render(request, 'setting_org/importSetting.html')

@requiredLogin
def exportExcelTemplate(request: HttpRequest):
    try:
        orgSettings = OrgSetting.objects.filter(isActive = True)
        excel = CreateExcelTemplateSetting("Org Setting", "app_organization", orgSettings)
        if excel:
            # สร้าง response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="org-setting.xlsx"'
            # save workbook ลง response
            excel.save(response)
            return response
        messages.error(request, "excel template data not found")
        return HttpResponseRedirect(reverse('indexSettingOrg'))
    except Exception as e:
        print(e)
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse('indexSettingOrg'))