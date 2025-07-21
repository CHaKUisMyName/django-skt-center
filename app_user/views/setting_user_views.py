import json
import tempfile
from typing import List
from bson import ObjectId
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.contrib import messages
from django.utils import timezone
from openpyxl import Workbook, load_workbook, styles
from openpyxl.utils import get_column_letter

from app_system_setting.models import SystemApp, SystemMenu
from app_user.models.user import User
from app_user.models.user_setting import UserSetting
from app_user.utils import isSettingUserAdmin, requiredLogin
from base_models.basemodel import UserSnapshot

@requiredLogin
def indexSettingUser(request: HttpRequest):
    userSettings = UserSetting.objects.filter(isActive = True)
    isUserAdmin = isSettingUserAdmin(request.currentUser.id)
    if not request.currentUser.isAdmin or isUserAdmin == False:
        messages.error(request, "Not Permission")
        return HttpResponseRedirect('/')
    context = {
        "userSettings": userSettings
    }
    return render(request, 'setting_user/indexSetting.html', context)

@requiredLogin
def addSettingUser(request: HttpRequest):
    response = HttpResponseRedirect(reverse('indexSettingUser'))
    isUserAdmin = isSettingUserAdmin(request.currentUser.id)
    if not request.currentUser.isAdmin or isUserAdmin == False:
        messages.error(request, "Not Permission")
        return HttpResponseRedirect('/')
    if request.method == "POST":
        try:
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            user = request.POST.get("user")
            if not user:
                messages.error(request, "User is required")
                return response
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            dupUser: UserSetting = UserSetting.objects.filter(user = user).first()
            if dupUser:
                messages.error(request, "Duplicate User")
                return response
            
            userSetting = UserSetting()
            userSetting.user = ObjectId(user)
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            userSetting.menus = menuList
            userSetting.isAdmin = isadmin
            userSetting.isActive = True
            currentUser: User = request.currentUser
            if currentUser:
                uCreate = UserSnapshot().UserToSnapshot(currentUser)
                if uCreate:
                    userSetting.createBy = uCreate
            userSetting.save()

            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response

    else:
        users: List[User] = User.objects.filter(isActive = True)
        app: SystemApp = SystemApp.objects.filter(name = "app_user").first()
        if not app:
            messages.error(request, "App not found")
            return response
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        selectedUsers = []
        userSettings: List[UserSetting] = UserSetting.objects.filter(isActive = True)
        if userSettings:
            selectedUsers = [us.user.id for us in userSettings]


        context = {
            "users": users,
            "menus": menus,
            "selectedUsers": selectedUsers
        }
        return render(request, 'setting_user/addSetting.html', context)
    
@requiredLogin
def editSettingUser(request: HttpRequest, id: str):
    response = HttpResponseRedirect(reverse('indexSettingUser'))
    isUserAdmin = isSettingUserAdmin(request.currentUser.id)
    if not request.currentUser.isAdmin or isUserAdmin == False:
        messages.error(request, "Not Permission")
        return HttpResponseRedirect('/')
    if request.method == "POST":
        try:
            sid = request.POST.get("id")
            if not id:
                messages.error(request, "Id is required")
                return response
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            
            userSetting = UserSetting.objects.filter(id = sid).first()
            if not userSetting:
                messages.error(request, "User Setting not found")
                return response
            
            userSetting.isAdmin = isadmin
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            userSetting.menus = menuList
            userSetting.updateDate = timezone.now()

            currentUser: User = request.currentUser
            if currentUser:
                uUpdate = UserSnapshot().UserToSnapshot(currentUser)
                if uUpdate:
                    userSetting.updateBy = uUpdate
            userSetting.save()

            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response
    else:
        users: List[User] = User.objects.filter(isActive = True)
        app: SystemApp = SystemApp.objects.filter(name = "app_user").first()
        if not app:
            messages.error(request, "App not found")
            return response
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        userSetting: UserSetting = UserSetting.objects.filter(id = id).first()
        if not userSetting:
            messages.error(request, "User Setting not found")
            return response
        
        menu_ids = [m.id for m in userSetting.menus]

        context = {
            "users": users,
            "menus": menus,
            "userSetting": userSetting,
            "menu_ids": menu_ids
        }
        return render(request, 'setting_user/editSetting.html',context)
    
@requiredLogin
def deleteSettingUser(request: HttpRequest, id: str):
    try:
        isUserAdmin = isSettingUserAdmin(request.currentUser.id)
        if not request.currentUser.isAdmin or isUserAdmin == False:
            return JsonResponse({'deleted': False, 'message': 'Not Permission'})
        if not request.method == "GET":
            return JsonResponse({'deleted': False, 'message': 'Method not allowed'})
        if not id:
            return JsonResponse({'deleted': False, 'message': 'Not found id'})
        
        userSetting: UserSetting = UserSetting.objects.filter(id = id).first()
        if not userSetting:
            return JsonResponse({'deleted': False, 'message': 'User Setting not found'})
        userSetting.isActive = False
        userSetting.updateDate = timezone.now()
        currentUser: User = request.currentUser
        if currentUser:
            uUpdate = UserSnapshot().UserToSnapshot(currentUser)
            if uUpdate:
                userSetting.updateBy = uUpdate
        userSetting.save()
        
        return JsonResponse({'deleted': True, 'message': 'Delete success'})
    except Exception as e:
        return JsonResponse({'deleted': False, 'message': str(e)})
    
@requiredLogin
def importSettingUser(request: HttpRequest):
    isUserAdmin = isSettingUserAdmin(request.currentUser.id)
    if not request.currentUser.isAdmin or isUserAdmin == False:
        messages.error(request, "Not Permission")
        return HttpResponseRedirect('/')
    if request.method == "POST":
        response = HttpResponseRedirect(reverse('indexSettingUser'))
        try:
            file = request.FILES.get("file")
            if not file:
                messages.error(request, "File is required")
                return HttpResponseRedirect(reverse('importSettingUser'))
            if not file.name.endswith(".xlsx"):
                messages.error(request, "Only .xlsx files are supported.")
                return HttpResponseRedirect(reverse('importSettingUser'))
            with tempfile.NamedTemporaryFile(delete=True, suffix=".xlsx") as temp_file:
                for chunk in file.chunks():
                    temp_file.write(chunk)
                
                # ✅ ข้อมูลถูกเขียนลงไฟล์ .xlsx บน disk จริง
                # flush() จะสั่งว่า: "เขียนเลยตอนนี้นะ!"
                temp_file.flush()
                # ย้าย “ตำแหน่ง pointer” ในไฟล์ — ว่าเราจะเริ่มอ่านหรือเขียนตรงไหน
                # หลังจากเขียนไฟล์เสร็จ pointer จะอยู่ "ท้ายไฟล์"
                # ถ้าจะเปิดอ่านต่อ (หรือใช้ file object เดิมอ่าน) ต้อง “ย้อนกลับไปที่ต้นไฟล์” → seek(0)
                temp_file.seek(0)

                wb = load_workbook(temp_file.name)
                ws = wb.active
                usSettings = []
                header_row = [cell.value for cell in ws[2]]
                menu_names = header_row[3:]  # คอลัมน์เมนูเริ่มจาก index 4
                # ดึง SystemMenu ทั้งหมด
                system_menus = {menu.name: menu for menu in SystemMenu.objects(name__in=menu_names)}
                inserted_codes = set()
                rowCount = 3
                for row in ws.iter_rows(min_row=3):
                    code = str(row[0].value).strip() if row[0].value else None
                    if not code:
                        continue  # skip ถ้าไม่มี code 

                    if code in inserted_codes:
                        messages.error(request, f"Duplicate code: {code} in row ({rowCount})")# skip ถ้า code ซ้ำในไฟล์เดิม
                        return HttpResponseRedirect(reverse('importSettingUser'))
                    
                    user:User = User.objects.filter(code=code).first()
                    if not user:
                        messages.error(request, f"User not found in row ({rowCount}) with code: {code}")
                        return HttpResponseRedirect(reverse('importSettingUser'))
                    
                    checkUsSetting: UserSetting = UserSetting.objects.filter(user = user).first()
                    if checkUsSetting:
                        if checkUsSetting.isActive == True:
                            # -- ถ้ายัง active ให้ show error
                            messages.error(request, f"Duplicate code: {code} in row ({rowCount})")
                            return HttpResponseRedirect(reverse('importSettingUser'))
                        else:
                            # -- ถ้ามีข้อมูลใน db แล้ว ไม่ active false ให้ลบข้อมูลออกจาก db แล้ว save ของใหม่จาก excel ไปเลย
                            checkUsSetting.delete()

                    selected_menu_names = []
                    for i, cell in enumerate(row[3:], start=3):
                        if str(cell.value).strip() == "1":
                            selected_menu_names.append(header_row[i])

                    if not selected_menu_names:
                        continue  # ข้ามถ้าไม่ได้เลือกเมนู

                    menu_refs = [system_menus[name] for name in selected_menu_names if name in system_menus]

                    # สร้าง UserSetting
                    setting: UserSetting = UserSetting()
                    setting.user=user
                    setting.menus=menu_refs
                    setting.isActive=True
                    setting.isAdmin=False
                    setting.note=''
                    currentUser: User = request.currentUser
                    if currentUser:
                        uCreate = UserSnapshot().UserToSnapshot(currentUser)
                        if uCreate:
                            setting.createBy=uCreate
                    setting.createDate=timezone.now()

                    
                    usSettings.append(setting)
                    inserted_codes.add(code)
                    rowCount += 1
                # print(usSettings)

                # MongoEngine สามารถ insert หลาย document พร้อมกัน
                # ทำงานเร็วกว่า for setting in usSettings: setting.save()
                UserSetting.objects.insert(usSettings)
            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return HttpResponseRedirect(reverse('importSettingUser'))
    else:
        return render(request, 'setting_user/importSetting.html')
    
@requiredLogin
def exportExcelTemplate(request: HttpRequest):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "User Setting"
        # ใส่ header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        ws.cell(row=1, column=1).value = "Employee"
        ws.cell(row=1, column=1).alignment = styles.Alignment(horizontal='center')
        app: SystemApp = SystemApp.objects.filter(name = "app_user").first()
        header = ["code", "fNameEN", "lNameEN"]
        if app:
            menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
            if menus:
                ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=4+(len(menus)-1))
                ws.cell(row=1, column=4).value = "Menu"
                ws.cell(row=1, column=4).alignment = styles.Alignment(horizontal='center')
                # -- size menu column
                for col_idx in range(4, (4+len(menus))):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                for m in menus:
                    header.append(m.name)
        ws.append(header)
        
        # ใส่ข้อมูล
        user: list[User] = User.objects.filter(isActive = True)
        if user:
            dupUser = []
            usSetting = UserSetting.objects.filter(isActive = True)
            if usSetting:
                dupUser = [us.user.id for us in usSetting]
            for u in user:
                if not u.id in dupUser:
                    ws.append([u.code, u.fNameEN, u.lNameEN])

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            
            for cell in row:
                if cell.row in [1,2]:
                    cell.alignment = styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = styles.Font(bold=True)
                    cell.border = styles.Border(
                        left=styles.Side(style='thin'), 
                        right=styles.Side(style='thin'),
                        top=styles.Side(style='thin'), 
                        bottom=styles.Side(style='thin')
                    )
                else:
                    cell.alignment = styles.Alignment(wrap_text=True)
                    cell.border = styles.Border(
                        left=styles.Side(style='thin'), 
                        right=styles.Side(style='thin'),
                        top=styles.Side(style='thin'), 
                        bottom=styles.Side(style='thin')
                    )
        # -- size name column
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # สร้าง response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="export.xlsx"'

        # save workbook ลง response
        wb.save(response)

        return response
    except Exception as e:
        print(e)
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse('indexSettingUser'))