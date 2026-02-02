from collections import defaultdict
import datetime
import json
from typing import List
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.contrib import messages
from django.urls import reverse
from openpyxl import Workbook, styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from app_user.models.opd import BudgetEmpType, BudgetOpd, OPDRecord, OpdType, OptionOpd, SpecialBudgetOpd
from app_user.models.user import EmpNation, User, UserStatus
from mongoengine.queryset.visitor import Q
import pytz
from django.utils import timezone

from app_user.utils import HasUsPermission, requiredLogin
from utilities.utility import findDeptUser

@requiredLogin
def index(request: HttpRequest):
    hasPermission = HasUsPermission(id = str(request.currentUser.id), menu = "Opd-Admin")
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    try:
        budgetTH: BudgetOpd = BudgetOpd.objects.filter(isActive = True, type = BudgetEmpType.Thai, status = True).first()
        budgetFore: BudgetOpd = BudgetOpd.objects.filter(isActive = True, type = BudgetEmpType.Foreigner, status = True).first()
        summary = summaryBudget(budgetTH, budgetFore)

        cur_date = timezone.now()
        start_date = datetime.datetime(cur_date.year, cur_date.month, 1, tzinfo=pytz.UTC)
        end_date = datetime.datetime(cur_date.year + 1, cur_date.month, 1, tzinfo=pytz.UTC)

        creditOption: OptionOpd = OptionOpd.objects.filter(name = "Credit").first()
        findRecordCreditOption: List[OPDRecord] = OPDRecord.objects.filter(
            isActive = True,
            inputDate__gte = start_date,
            inputDate__lt = end_date,
            disbursements__refOption = creditOption
        )
        print(f"len() : {len(findRecordCreditOption)}")
        sumCreditAmount = 0
        for item in findRecordCreditOption:
            for disb in item.disbursements:
                sumCreditAmount += disb.amount
        
        context = {
            'summary': summary,
            'summaryFormat': f"{summary:,.2f}",
            'sumCreditAmount': f"{sumCreditAmount:,.2f}",
            "curMonth": f"{cur_date.month}/{cur_date.year}",
            'budgetTH': f"{budgetTH.budget:,.2f}" if budgetTH else 0,
            'budgetFore': f"{budgetFore.budget:,.2f}" if budgetFore else 0
        }
        return render(request, 'opd/report_opd.html', context)
    except Exception as e:
        messages.error(request, str(e))
        return HttpResponseRedirect('/')
    
def summaryBudget(budgetTH, budgetFore):
    try:
        totalBudget = 0

        thaiEmp = User.objects.filter(isActive = True, nation = EmpNation.Thai.value, status = UserStatus.Hire.value).only('id')
        foreignerEmp = User.objects.filter(isActive = True, nation__ne = EmpNation.Thai.value, status = UserStatus.Hire.value).only('id')

        # -- หา special budget แยก ไทย ต่างประเทศ
        specialBudTH = SpecialBudgetOpd.objects.filter(isActive = True, employee__in = thaiEmp)
        specialBudFore = SpecialBudgetOpd.objects.filter(isActive = True, employee__in = foreignerEmp)

        totalSpecialFore = 0
        specialBudForeCount = 0
        if specialBudFore:
            specialBudForeCount = specialBudFore.count()
            for special in specialBudFore:
                totalSpecialFore += special.specialBudget
        
        totalSpecialTH = 0
        specialBudTHCount = 0
        if specialBudTH:
            specialBudTHCount = specialBudTH.count()
            for special in specialBudTH:
                totalSpecialTH += special.specialBudget
        
        totalBudgetFore = (foreignerEmp.count() - specialBudForeCount) * budgetFore.budget
        totalBudgetTh = (thaiEmp.count() - specialBudTHCount) * budgetTH.budget
        totalBudget = totalBudgetFore + totalBudgetTh + totalSpecialFore + totalSpecialTH
        
        return totalBudget
    except Exception as e:
        print(e)
        return 0

@requiredLogin
def pieChart(request: HttpRequest):
    if not request.method == "POST":
            return JsonResponse({'success': False, 'data': [0,0,0], 'message': 'Method not allowed'})
    try:
        result = [0,0,0]
        dateNow = timezone.now()
        start_date = datetime.datetime(dateNow.year, 1, 1, tzinfo=pytz.UTC)
        body = json.loads(request.body.decode('utf-8'))  # อ่าน JSON body
        budget = body.get('budget')
        if not budget:
            return JsonResponse({'success': False, 'data': result, 'message': 'Budget is required'})
        
        opdRecord: List[OPDRecord] = OPDRecord.objects.filter(
            isActive = True,
            inputDate__gte = start_date,
            inputDate__lt = dateNow,
        )

        diff = 0
        sumEmp = 0
        sumFamily = 0

        if opdRecord:
            for opd in opdRecord:
                for disb in opd.disbursements:
                    if disb.type == OpdType.Employee:
                        sumEmp += disb.amount
                    if disb.type == OpdType.Family:
                        sumFamily += disb.amount
        
        diff = float(budget) -  (sumEmp + sumFamily)
        result[0] = sumEmp
        result[1] = sumFamily
        result[2] = diff

        print(f"""result : {result} """)

        return JsonResponse({'success': True, 'data': result, 'message': 'Success'})
    except Exception as e:
        print(e)
        return JsonResponse({'success': True, 'data': [0,0,0], 'message': 'Success'})
    
@requiredLogin
def exportReportOPD(request: HttpRequest):
    try:
        users: List[User] = User.objects.filter(isActive = True, status = UserStatus.Hire.value).order_by('code')
        if not users:
            messages.error(request, "User not found")
            return HttpResponseRedirect(reverse('indexReportOpd'))
        # allOption: List[OptionOpd] = OptionOpd.objects.filter(isActive = True, status = True).order_by('name')
        # if not allOption:
        #     messages.error(request, "Option not found")
        #     return HttpResponseRedirect(reverse('indexReportOpd'))
        start_date = datetime.datetime(timezone.now().year, 1, 1, tzinfo=pytz.UTC)
        end_date = datetime.datetime(timezone.now().year + 1, 1, 1, tzinfo=pytz.UTC)
        opdRecord: List[OPDRecord] = OPDRecord.objects.filter(
            isActive = True,
            inputDate__gte = start_date,
            inputDate__lt = end_date,
            )
        opd_by_emp = defaultdict(list)
        for rec in opdRecord:
            if rec.employee:
                # {'a': [1, 2]}
                opd_by_emp[str(rec.employee.id)].append(rec)

        buggets: List[BudgetOpd] = BudgetOpd.objects.filter(isActive = True, status = True)
        budTH = buggets.filter(type = BudgetEmpType.Thai).first()
        budFore = buggets.filter(type = BudgetEmpType.Foreigner).first()
        
        specialBudgets: List[SpecialBudgetOpd] = SpecialBudgetOpd.objects.filter(isActive = True)
        
        # opEmp = allOption.filter(useTypes__in=[OpdType.Employee])
        # opFm  = allOption.filter(useTypes__in=[OpdType.Family])
        opEmp = OptionOpd.objects.filter(isActive = True, status = True, useTypes__in=[OpdType.Employee]).order_by('name')
        opFm  = OptionOpd.objects.filter(isActive = True, status = True, useTypes__in=[OpdType.Family]).order_by('name')
        
        optionKeyEmp = [
            f"emp-{op.name}"
            for op in opEmp
        ]
        optionKeyFm = [
            f"fm-{op.name}"
            for op in opFm
        ]

        listUserObjOPD = []
        for user in users:
            dept = findDeptUser(user)
            deptStr = ''
            if len(dept) > 0:
                for de in dept:
                    deptStr += f"{de['nameEN']}, "
            usObj = {
                'code': user.code,
                'fullName': f"{user.fNameEN} {user.lNameEN}",
                'dept': deptStr,
            }
            for key in optionKeyEmp:
                usObj[key] = 0
            for key in optionKeyFm:
                usObj[key] = 0

            emp_records = opd_by_emp.get(str(user.id), [])
            for rec in emp_records:
                for disb in rec.disbursements:
                    if disb is None or disb.refOption is None or disb.amount is None:
                        continue

                    optionName = disb.refOption.name
                    amount = disb.amount or 0

                    if disb.type == OpdType.Employee:
                        key = f"emp-{optionName}"
                    elif disb.type == OpdType.Family:
                        key = f"fm-{optionName}"
                    else:
                        continue
                    # กันกรณี key ไม่มี (edge case)
                    if key not in usObj:
                        usObj[key] = 0

                    usObj[key] += amount

            findUsinSpecial = specialBudgets.filter(employee = user).first()
            budget = 0

            if findUsinSpecial and findUsinSpecial.specialBudget is not None:
                budget = findUsinSpecial.specialBudget
            else:
                if user.nation == EmpNation.Thai and budTH and budTH.budget is not None:
                    budget = budTH.budget
                elif user.nation != EmpNation.Thai and budFore and budFore.budget is not None:
                    budget = budFore.budget

            usObj['budget'] = budget


            empTotal = 0
            famTotal = 0
            for k,v in usObj.items():
                if k.startswith("emp-"):
                    empTotal += v
                elif k.startswith("fm-"):
                    famTotal += v
            usObj['empTotal'] = empTotal
            usObj['famTotal'] = famTotal
            usObj['balance'] = usObj['budget'] - (empTotal + famTotal)
            # usObj : {'code': '', 'fullName': '', 'dept': '', ', 'emp-Credit': 0, 'emp-Dental': 0, 'emp-General': 0, 'emp-Glasses': 0, 'emp-Sport': 0, 'fm-Credit': 0, 'fm-Dental': 0, 'fm-General': 0, 'fm-Glasses': 0, 'budget': 0, 'empTotal': 0, 'famTotal': 0, 'balance': 0}
            listUserObjOPD.append(usObj)
        excel = createExcelReportOPD(listUserObjOPD, opEmp, opFm, optionKeyEmp, optionKeyFm)
        if excel:
            # สร้าง response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="opd-report.xlsx"'
            # save workbook ลง response
            excel.save(response)
            return response
        messages.error(request, "excel template data not found")
        return HttpResponseRedirect(reverse('indexReportOpd'))
    except Exception as e:
        print(e)
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse('indexReportOpd'))
    
def createExcelReportOPD(users, opEmp, opFm, optionKeyEmp, optionKeyFm):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "OPD-Report"

        border = styles.Border(
            left=styles.Side(style='thin'),
            right=styles.Side(style='thin'),
            top=styles.Side(style='thin'),
            bottom=styles.Side(style='thin')
        )

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        # ---------- COLORS ----------
        colorBlue = css_hex_to_argb("#c1c3f9")
        colorOrange = css_hex_to_argb("#f8e4c8")
        colorGreen = css_hex_to_argb("#b1fbd9")
        colorYellow = css_hex_to_argb("#f7ef3a")
        colorRed = css_hex_to_argb("#f80000")

        fill_blue = fill(colorBlue)
        fill_orange = fill(colorOrange)
        fill_green = fill(colorGreen)
        fill_yellow = fill(colorYellow)
        fill_red = fill(colorRed)


        # ---------- TITLE ----------
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.cell(row=1, column=1, value="OPD Report")
        ws.cell(row=1, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=1).font = styles.Font(bold=True, size=28)
        ws.row_dimensions[1].height = 77

        # ---------- FIXED HEADER ----------
        headers = ["No.", "Code", "Name", "Dept"]
        widths = [8, 20, 35, 30]

        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            ws.merge_cells(start_row=4, start_column=i, end_row=5, end_column=i)
            cell = ws.cell(row=4, column=i, value=h)
            cell.alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = styles.Font(bold=True, size=16)
            cell.border = border
            ws.column_dimensions[get_column_letter(i)].width = w


        # ---------- EMPLOYEE HEADER ----------
        emp_col_start = 5
        emp_col_end = emp_col_start + len(optionKeyEmp) - 1

        ws.merge_cells(
            start_row=4,
            start_column=emp_col_start,
            end_row=4,
            end_column=emp_col_end
        )

        ws.cell(row=4, column=emp_col_start, value="Employee")
        ws.cell(row=4, column=emp_col_start).alignment = styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=emp_col_start).font = styles.Font(bold=True, size=16)
        
        for col in range(emp_col_start, emp_col_end + 1):
            ws.cell(row=4, column=col).border = border
            ws.cell(row=4, column=col).fill = fill_blue

        
        print("test")

        col_idx = emp_col_start
        for op in opEmp:
            c = ws.cell(row=5, column=col_idx, value=op.name)
            c.alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.font = styles.Font(bold=True, size=16)
            c.border = border
            c.fill = fill_blue
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
            col_idx += 1
        

        # ---------- FAMILY HEADER ----------
        fm_col_start = col_idx
        fm_col_end = fm_col_start + len(optionKeyFm) - 1

        ws.merge_cells(
            start_row=4,
            start_column=fm_col_start,
            end_row=4,
            end_column=fm_col_end
        )

        ws.cell(row=4, column=fm_col_start, value="Family")
        ws.cell(row=4, column=fm_col_start).alignment = styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=fm_col_start).font = styles.Font(bold=True, size=16)

        for col in range(fm_col_start, fm_col_end + 1):
            ws.cell(row=4, column=col).border = border
            ws.cell(row=4, column=col).fill = fill_orange

        for op in opFm:
            c = ws.cell(row=5, column=col_idx, value=op.name)
            c.alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.font = styles.Font(bold=True, size=16)
            c.border = border
            c.fill = fill_orange
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
            col_idx += 1

        # ---------- TOTAL AMOUNT ----------
        total_start = col_idx
        ws.merge_cells(
            start_row=4,
            start_column=total_start,
            end_row=4,
            end_column=total_start + 3
        )

        ws.cell(row=4, column=total_start, value="Total Amount")
        ws.cell(row=4, column=total_start).alignment = styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=total_start).font = styles.Font(bold=True, size=16)

        for col in range(total_start, total_start + 4):
            ws.cell(row=4, column=col).border = border
            ws.cell(row=4, column=col).fill = fill_green
            ws.column_dimensions[get_column_letter(col)].width = 20

        totals = ["Budget", "Employee", "Family", "Balance"]
        for i, t in enumerate(totals):
            c = ws.cell(row=5, column=total_start + i, value=t)
            c.alignment = styles.Alignment(horizontal="center", vertical="center")
            c.font = styles.Font(bold=True, size=16)
            c.fill = fill_green
            c.border = border

        # ---------- DATA ----------
        start_row = 6
        
        for idx, user in enumerate(users, start=1):
            r = start_row + idx - 1
            row = [
                idx,
                user["code"],
                user["fullName"],
                user["dept"],
            ]

            for k in optionKeyEmp:
                row.append(user.get(k, 0))
            for k in optionKeyFm:
                row.append(user.get(k, 0))
            # kuy = f"=SUM({ws.cell(row=r, column=ws.max_column-3).coordinate}:{ws.cell(row=r, column=ws.max_column-1).coordinate})"
            # print(kuy)
            row.extend([
                user["budget"],
                user["empTotal"],
                user["famTotal"],
                user["balance"]
                # kuy
            ])

            ws.append(row)
            ws.row_dimensions[r].height = 25
            cell_index = 1
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.alignment = styles.Alignment(wrap_text=True)
                if c != 1 and c != 2 and c != 3 and c != 4:
                    cell.number_format = '#,##0.00'
                
                if cell_index < ws.max_column and cell_index >= total_start:
                    cell.fill = fill_green
                    cell.font = styles.Font(bold=True, size=16)
                elif cell_index <= fm_col_end and cell_index >= fm_col_start:
                    cell.fill = fill_orange
                    cell.font = styles.Font(bold=True, size=16)
                elif cell_index <= emp_col_end and cell_index >= emp_col_start:
                    cell.fill = fill_blue
                    cell.font = styles.Font(bold=True, size=16)

                cell_index += 1
                
            # balance color
            bal_cell = ws.cell(row=r, column=ws.max_column)
            bal_cell.font = styles.Font(bold=True, size=16)
            bal_cell.fill = fill_green
            bal_cell.number_format = '#,##0.00'
            balance = user.get("balance", 0) or 0
            

            if balance <= 0:
                bal_cell.font = styles.Font(bold=True, size=16, color='FF0000')

        # ---------- STAMP DATE ROW ----------
        curDate = timezone.now()
        # -- to dd/mm/yyyy
        dateStr = curDate.strftime("%d/%m/%Y")
        cell_stamp_label = ws.cell(row=2, column=ws.max_column-1)
        cell_stamp_label.value = "Stamp Date"
        cell_stamp_label.font = styles.Font(bold=True, size=16)

        cell_stamp_date = ws.cell(row=2, column=ws.max_column)
        cell_stamp_date.value = dateStr
        cell_stamp_date.font = styles.Font(size=16)


        # ---------- SUMMARY ROW ----------
        summary_row = ws.max_row + 1
        sum_cell_balance = ws.cell(row=summary_row, column=ws.max_column)
        # sum_cell_balance.value = sumBalance
        sum_cell_balance.value = f"=SUM({ws.cell(row=6, column=ws.max_column).coordinate}:{ws.cell(row=ws.max_row-1, column=ws.max_column).coordinate})"
        sum_cell_balance.font = styles.Font(bold=True, size=16)
        sum_cell_balance.fill = fill_yellow
        sum_cell_balance.number_format = '#,##0.00'
        sum_cell_balance.border = border

        sum_cell_fam = ws.cell(row=summary_row, column=ws.max_column-1)
        # sum_cell_fam.value = sumFm
        sum_cell_fam.value = f"=SUM({ws.cell(row=6, column=ws.max_column-1).coordinate}:{ws.cell(row=ws.max_row-1, column=ws.max_column-1).coordinate})"
        sum_cell_fam.font = styles.Font(bold=True, size=16)
        sum_cell_fam.fill = fill_yellow
        sum_cell_fam.number_format = '#,##0.00'
        sum_cell_fam.border = border

        sum_cell_emp = ws.cell(row=summary_row, column=ws.max_column-2)
        # sum_cell_emp.value = sumEmp
        sum_cell_emp.value = f"=SUM({ws.cell(row=6, column=ws.max_column-2).coordinate}:{ws.cell(row=ws.max_row-1, column=ws.max_column-2).coordinate})"
        sum_cell_emp.font = styles.Font(bold=True, size=16)
        sum_cell_emp.fill = fill_yellow
        sum_cell_emp.number_format = '#,##0.00'
        sum_cell_emp.border = border


        sum_cell_budg = ws.cell(row=summary_row, column=ws.max_column-3)
        # sum_cell_budg.value = sumBudget
        sum_cell_budg.value = f"=SUM({ws.cell(row=6, column=ws.max_column-3).coordinate}:{ws.cell(row=ws.max_row-1, column=ws.max_column-3).coordinate})"
        sum_cell_budg.font = styles.Font(bold=True, size=16)
        sum_cell_budg.fill = fill_yellow
        sum_cell_budg.number_format = '#,##0.00'
        sum_cell_budg.border = border


        ws.cell(row=summary_row, column=ws.max_column-5).border = border
        ws.merge_cells(start_row=summary_row, start_column=ws.max_column-5, end_row=summary_row, end_column=ws.max_column-4)
        sum_merge_cell = ws.cell(row=summary_row, column=ws.max_column-5, value="Total")
        sum_merge_cell.font = styles.Font(bold=True, size=16)
        sum_merge_cell.fill = fill_yellow
        sum_merge_cell.alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # ---------- PERCENTAGE ROW ----------
        percent_row = ws.max_row + 1

        per_cell_balance = ws.cell(row=percent_row, column=ws.max_column)
        # per_cell_balance.value = sumBalance/sumBudget
        per_cell_balance.value = f"={ws.cell(row=percent_row-1, column=ws.max_column).coordinate}/{ws.cell(row=percent_row-1, column=ws.max_column-3).coordinate}"
        per_cell_balance.font = styles.Font(bold=True, size=16)
        per_cell_balance.fill = fill_yellow
        per_cell_balance.number_format = '0.00%'
        per_cell_balance.border = border


        per_cell_fm = ws.cell(row=percent_row, column=ws.max_column-1)
        # per_cell_fm.value = sumFm/sumBudget
        per_cell_fm.value = f"={ws.cell(row=percent_row-1, column=ws.max_column-1).coordinate}/{ws.cell(row=percent_row-1, column=ws.max_column-3).coordinate}"
        per_cell_fm.font = styles.Font(bold=True, size=16)
        per_cell_fm.fill = fill_yellow
        per_cell_fm.number_format = '0.00%'
        per_cell_fm.border = border

        per_cell_emp = ws.cell(row=percent_row, column=ws.max_column-2)
        # per_cell_emp.value = sumEmp/sumBudget
        per_cell_emp.value = f"={ws.cell(row=percent_row-1, column=ws.max_column-2).coordinate}/{ws.cell(row=percent_row-1, column=ws.max_column-3).coordinate}"
        per_cell_emp.font = styles.Font(bold=True, size=16)
        per_cell_emp.fill = fill_yellow
        per_cell_emp.number_format = '0.00%'
        per_cell_emp.border = border


        per_cell_emp = ws.cell(row=percent_row, column=ws.max_column-3)
        per_cell_emp.font = styles.Font(bold=True, size=16)
        per_cell_emp.fill = fill_yellow
        per_cell_emp.border = border


        ws.cell(row=percent_row, column=ws.max_column-5).border = border
        # ws.cell(row=percent_row, column=ws.max_column-5).border = border
        ws.merge_cells(start_row=percent_row, start_column=ws.max_column-5, end_row=percent_row, end_column=ws.max_column-4)
        per_merge_cell = ws.cell(row=percent_row, column=ws.max_column-5, value="Percentage")
        per_merge_cell.font = styles.Font(bold=True, size=16)
        per_merge_cell.fill = fill_yellow
        per_merge_cell.alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

        return wb

    except Exception as e:
        print(e)
        return None
    
def css_hex_to_argb(css_hex):
    css_hex = css_hex.lstrip('#')
    if len(css_hex) == 8:  # RRGGBBAA
        return css_hex[6:8] + css_hex[0:6]
    elif len(css_hex) == 6:
        return "FF" + css_hex
    else:
        raise ValueError("Invalid hex color")
