import tempfile
from typing import List
from bson import ObjectId
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from openpyxl import load_workbook
from app_safety_system.models.safety_setting import SafetySetting
from app_safety_system.utils import HasSftPermission
from app_system_setting.models import SystemApp, SystemMenu
from app_user.models.user import User, UserStatus
from app_user.utils import requiredLogin
from django.contrib import messages
from django.utils import timezone

from base_models.basemodel import UserSnapshot
from utilities.utility import CreateExcelTemplateSetting


@requiredLogin
def indexSettingSafety(request: HttpRequest):
    hasPermission = HasSftPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    safetySettings = SafetySetting.objects.filter(isActive = True)
    context = {
        'safetySettings': safetySettings
    }
    return render(request, 'setting_safety/indexSetting.html', context)

@requiredLogin
def addSettingSafety(request: HttpRequest):
    hasPermission = HasSftPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    response = HttpResponseRedirect(reverse('indexSettingSafety'))
    if request.method == "POST":
        try:
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            user = request.POST.get("user")
            if not user:
                messages.error(request, "User is required")
                return response
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            dupUser: SafetySetting = SafetySetting.objects.filter(user = user).first()
            if dupUser:
                messages.error(request, "Duplicate User")
                return response
            newSetting = SafetySetting()
            newSetting.user = ObjectId(user)
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            newSetting.menus = menuList
            newSetting.isAdmin = isadmin
            newSetting.isActive = True
            currentUser: User = request.currentUser
            if currentUser:
                uCreate = UserSnapshot().UserToSnapshot(currentUser)
                if uCreate:
                    newSetting.uCreate = uCreate
            newSetting.save()

            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response
    else:
        users: List[User] = User.objects.filter(isActive = True, status = UserStatus.Hire.value).order_by('code')
        app: SystemApp = SystemApp.objects.filter(name = "app_safety_system").first()
        if not app:
            messages.error(request, "App not found")
            return response
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        selectedUsers = []
        safetySettings: List[SafetySetting] = SafetySetting.objects.filter(isActive = True)
        if safetySettings:
            for setting in safetySettings:
                selectedUsers.append(setting.user.id)
        context = {
            'users': users,
            'menus': menus,
            'selectedUsers': selectedUsers
        }
        return render(request, 'setting_safety/addSetting.html', context)
    
@requiredLogin
def editSettingSafety(request: HttpRequest, id: str):
    hasPermission = HasSftPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    response = HttpResponseRedirect(reverse('indexSettingSafety'))
    if request.method == "POST":
        try:
            sid = request.POST.get("id")
            if not id:
                messages.error(request, "Id is required")
                return response
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            setting: SafetySetting = SafetySetting.objects.filter(id = ObjectId(sid)).first()
            if not setting:
                messages.error(request, "Setting not found")
                return response
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            setting.menus = menuList
            setting.isAdmin = isadmin
            setting.updateDate = timezone.now()
            currentUser: User = request.currentUser
            if currentUser:
                uUpdate = UserSnapshot().UserToSnapshot(currentUser)
                if uUpdate:
                    setting.updateBy = uUpdate
            setting.save()

            messages.success(request, 'Update Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response
    else:
        users: List[User] = User.objects.filter(isActive = True).order_by('code')
        setting: SafetySetting = SafetySetting.objects.filter(id = ObjectId(id)).first()
        if not setting:
            messages.error(request, "Setting not found")
            return response
        app: SystemApp = SystemApp.objects.filter(name = "app_safety_system").first()
        if not app:
            messages.error(request, "App not found")
            return response
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        
        menu_ids = [m.id for m in setting.menus]

        context = {
            "users": users,
            "menus": menus,
            'safetySetting': setting,
            "menu_ids": menu_ids
        }
        return render(request, 'setting_safety/editSetting.html', context)
    
@requiredLogin
def deleteSettingSafety(request: HttpRequest, id: str):
    try:
        hasPermission = HasSftPermission(id = str(request.currentUser.id), checkAdmin = True)
        if not request.currentUser.isAdmin:
            if hasPermission == False:
                return JsonResponse({'success': False, 'message': 'Not Permission'})
        if not request.method == 'GET':
            return JsonResponse({'success': False, 'message': 'Method not allowed'})
        if not id:
            return JsonResponse({'success': False, 'message': 'Id is required'})
        setting: SafetySetting = SafetySetting.objects.filter(id = ObjectId(id)).first()
        if not setting:
            return JsonResponse({'success': False, 'message': 'Setting not found'})
        setting.isActive = False
        setting.updateDate = timezone.now()
        currentUser: User = request.currentUser
        if currentUser:
            uUpdate = UserSnapshot().UserToSnapshot(currentUser)
            if uUpdate:
                setting.updateBy = uUpdate
        setting.save()
        return JsonResponse({'success': True, 'message': 'Success'})
    except Exception as e:
        print(e)
        return JsonResponse({'success': False, 'message': str(e)})
    
@requiredLogin
def importSettingSafety(request: HttpRequest):
    hasPermission = HasSftPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    response = HttpResponseRedirect(reverse('indexSettingSafety'))
    if request.method == "POST":
        try:
            file = request.FILES.get("file")
            if not file:
                messages.error(request, "File is required")
                return HttpResponseRedirect(reverse('importSettingSafety'))
            if not file.name.endswith(".xlsx"):
                messages.error(request, "Only .xlsx files are supported.")
                return HttpResponseRedirect(reverse('importSettingSafety'))
            with tempfile.NamedTemporaryFile(delete=True, suffix=".xlsx") as temp_file:
                for chunk in file.chunks():
                    temp_file.write(chunk)
                    # ✅ ข้อมูลถูกเขียนลงไฟล์ .xlsx บน disk จริง
                    # flush() จะสั่งว่า: "เขียนเลยตอนนี้นะ!"
                    temp_file.flush()
                    # ย้าย “ตำแหน่ง pointer” ในไฟล์ — ว่าเราจะเริ่มอ่านหรือเขียนตรงไหน
                    # หลังจากเขียนไฟล์เสร็จ pointer จะอยู่ "ท้ายไฟล์"
                    # ถ้าจะเปิดอ่านต่อ (หรือใช้ file object เดิมอ่าน) ต้อง “ย้อนกลับไปที่ต้นไฟล์” → seek(0)
                    temp_file.seek(0)

                    wb = load_workbook(temp_file.name)
                    ws = wb.active
                    usSettings = []
                    header_row = [cell.value for cell in ws[2]]
                    menu_names = header_row[3:]  # คอลัมน์เมนูเริ่มจาก index 4
                    # ดึง SystemMenu ทั้งหมด
                    system_menus = {menu.name: menu for menu in SystemMenu.objects(name__in=menu_names)}
                    inserted_codes = set()
                    rowCount = 3
                    for row in ws.iter_rows(min_row=3):
                        code = str(row[0].value).strip() if row[0].value else None
                        if not code:
                            continue  # skip ถ้าไม่มี code 

                        if code in inserted_codes:
                            messages.error(request, f"Duplicate code: {code} in row ({rowCount})")# skip ถ้า code ซ้ำในไฟล์เดิม
                            return HttpResponseRedirect(reverse('importSettingSafety'))
                        
                        user:User = User.objects.filter(code=code).first()
                        if not user:
                            messages.error(request, f"User not found in row ({rowCount}) with code: {code}")
                            return HttpResponseRedirect(reverse('importSettingSafety'))
                        
                        checkSftSetting: SafetySetting = SafetySetting.objects.filter(user = user).first()
                        if checkSftSetting:
                            if checkSftSetting.isActive == True:
                                # -- ถ้ายัง active ให้ show error
                                messages.error(request, f"Duplicate code: {code} in row ({rowCount})")
                                return HttpResponseRedirect(reverse('importSettingSafety'))
                            else:
                                # -- ถ้ามีข้อมูลใน db แล้ว ไม่ active false ให้ลบข้อมูลออกจาก db แล้ว save ของใหม่จาก excel ไปเลย
                                checkSftSetting.delete()

                        selected_menu_names = []
                        for i, cell in enumerate(row[3:], start=3):
                            if str(cell.value).strip() == "1":
                                selected_menu_names.append(header_row[i])

                        if not selected_menu_names:
                            continue  # ข้ามถ้าไม่ได้เลือกเมนู

                        menu_refs = [system_menus[name] for name in selected_menu_names if name in system_menus]

                        # สร้าง UserSetting
                        setting: SafetySetting = SafetySetting()
                        setting.user=user
                        setting.menus=menu_refs
                        setting.isActive=True
                        setting.isAdmin=False
                        setting.note=''
                        currentUser: User = request.currentUser
                        if currentUser:
                            uCreate = UserSnapshot().UserToSnapshot(currentUser)
                            if uCreate:
                                setting.createBy = uCreate
                        setting.createDate=timezone.now()
                        usSettings.append(setting)
                        inserted_codes.add(code)
                        rowCount += 1

                    SafetySetting.objects.insert(usSettings)
                    

            messages.success(request, "Import Success")
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return HttpResponseRedirect(reverse('importSettingSafety'))
    else:
        return render(request, 'setting_safety/importSetting.html')
    
@requiredLogin
def exportExcelTemplate(request: HttpRequest):
    try:
        safetySetting: List[SafetySetting] = SafetySetting.objects.filter(isActive = True)
        excel = CreateExcelTemplateSetting("Safety Setting", "app_safety_system", safetySetting)
        if excel:
            # สร้าง response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="safety-setting.xlsx"'
            # save workbook ลง response
            excel.save(response)
            return response
        messages.error(request, "excel template data not found")
        return HttpResponseRedirect(reverse('indexSettingSafety'))
    except Exception as e:
        print(e)
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse('indexSettingSafety'))