import tempfile
from typing import List
from bson import ObjectId
from django.contrib import messages
from django.http import HttpRequest, HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from app_system_setting.models import SystemApp, SystemMenu
from app_user.models.user import User, UserStatus
from app_user.utils import requiredLogin
from app_visitor.models.visitor_setting import VisitorSetting
from app_visitor.utils import HasVstPermission
from base_models.basemodel import UserSnapshot
from utilities.utility import CreateExcelTemplateSetting

@requiredLogin
def index(request: HttpRequest):
    vstSettings = VisitorSetting.objects.filter(isActive = True)
    hasPermission = HasVstPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    context = {
        "vstSettings": vstSettings,
    }
    return render(request, 'setting_visitor/index.html', context)

@requiredLogin
def add(request: HttpRequest):
    response = HttpResponseRedirect(reverse('indexSettingVst'))
    hasPermission = HasVstPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    if request.method =='POST':
        try:
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            user = request.POST.get("user")
            if not user:
                messages.error(request, "User is required")
                return response
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            dupUser: VisitorSetting = VisitorSetting.objects.filter(user = user).first()
            if dupUser:
                messages.error(request, "Duplicate User")
                return response
            
            vstSetting = VisitorSetting()
            vstSetting.user = ObjectId(user)
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            vstSetting.menus = menuList
            vstSetting.isAdmin = isadmin
            vstSetting.isActive = True
            currentUser: User = request.currentUser
            if currentUser:
                uCreate = UserSnapshot().UserToSnapshot(currentUser)
                if uCreate:
                    vstSetting.createBy = uCreate
            vstSetting.save()
            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response
    else:
        users: List[User] = User.objects.filter(isActive = True, status = UserStatus.Hire.value).order_by('code')
        app: SystemApp = SystemApp.objects.filter(name = "app_visitor").first()
        if not app:
            messages.error(request, "App not found")
            return HttpResponseRedirect(reverse('indexSettingVst'))
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        selectedUsers = []
        orgSettings: List[VisitorSetting] = VisitorSetting.objects.filter(isActive = True)
        if orgSettings:
            selectedUsers = [us.user.id for us in orgSettings]
        
        context = {
            "users": users,
            "menus": menus,
            "selectedUsers": selectedUsers
        }

        return render(request, 'setting_visitor/addSetting.html', context)
    
@requiredLogin
def edit(request: HttpRequest, id: str):
    response = HttpResponseRedirect(reverse('indexSettingVst'))
    hasPermission = HasVstPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    if request.method == "POST":
        try:
            vstId = request.POST.get("vstId")
            if not vstId:
                messages.error(request, "Id is required")
                return response
            menus = request.POST.getlist("menus")
            if not menus:
                messages.error(request, "Menu is required")
                return response
            vstSetting: VisitorSetting = VisitorSetting.objects.filter(id = vstId).first()
            if not vstSetting:
                messages.error(request, "Visitor Setting not found")
                return response
            isadmin = True if request.POST.get("isadmin") == 'on' else False
            vstSetting.isAdmin = isadmin
            menuList = []
            for menu in menus:
                menuList.append(ObjectId(menu))
            vstSetting.menus = menuList
            vstSetting.updateDate = timezone.now()
            currentUser: User = request.currentUser
            if currentUser:
                uUpdateUser = UserSnapshot().UserToSnapshot(currentUser)
                if uUpdateUser:
                    vstSetting.updateBy = uUpdateUser
            vstSetting.save()

            messages.success(request, 'Save Success')
            return response
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return response
    else:
        if not id:
            messages.error(request, "Id is required")
            return response
        users: List[User] = User.objects.filter(isActive = True).order_by('code')
        app: SystemApp = SystemApp.objects.filter(name = "app_visitor").first()
        if not app:
            messages.error(request, "App not found")
            return response
        menus: List[SystemMenu] = SystemMenu.objects.filter(isActive = True, app = app.id)
        if not menus:
            messages.error(request, "Menu not found")
            return response
        vstSetting: VisitorSetting = VisitorSetting.objects.filter(id = id).first()
        if not vstSetting:
            messages.error(request, "Visitor Setting not found")
            return response
        menu_ids = [m.id for m in vstSetting.menus]
        context = {
            "users": users,
            "menus": menus,
            "vstSetting": vstSetting,
            "menu_ids": menu_ids
        }
        return render(request, 'setting_visitor/editSetting.html', context)
    
@requiredLogin
def delete(request: HttpRequest, id: str):
    try:
        if not request.method == "GET":
            return JsonResponse({'deleted': False, 'message': 'Method not allowed'})
        hasPermission = HasVstPermission(id = str(request.currentUser.id), checkAdmin = True)
        if not request.currentUser.isAdmin:
            if hasPermission == False:
                return JsonResponse({'deleted': False, 'message': 'Not Permission'})
        if not id:
            return JsonResponse({'deleted': False, 'message': 'Not found id'})
        vstSetting: VisitorSetting = VisitorSetting.objects.filter(id = id).first()
        if not vstSetting:
            return JsonResponse({'deleted': False, 'message': 'Visitor Setting not found'})
        vstSetting.isActive = False
        vstSetting.updateDate = timezone.now()
        currentUser: User = request.currentUser
        if currentUser:
            uUpdate = UserSnapshot().UserToSnapshot(currentUser)
            if uUpdate:
                vstSetting.updateBy = uUpdate
        vstSetting.save()

        return JsonResponse({'deleted': True, 'message': 'Delete success'})
    except Exception as e:
        return JsonResponse({'deleted': False, 'message': str(e)})
    
@requiredLogin
def importSettingVst(request: HttpRequest):
    hasPermission = HasVstPermission(id = str(request.currentUser.id), checkAdmin = True)
    if not request.currentUser.isAdmin:
        if hasPermission == False:
            messages.error(request, "Not Permission")
            return HttpResponseRedirect('/')
    if request.method == "POST":
        try:
            file = request.FILES.get("file")
            if not file:
                messages.error(request, "File is required")
                return HttpResponseRedirect(reverse('indexSettingVst'))
            if not file.name.endswith(".xlsx"):
                messages.error(request, "Only .xlsx files are supported.")
                return HttpResponseRedirect(reverse('indexSettingVst'))
            with tempfile.NamedTemporaryFile(delete=True, suffix=".xlsx") as temp_file:
                for chunk in file.chunks():
                    temp_file.write(chunk)
                # ✅ ข้อมูลถูกเขียนลงไฟล์ .xlsx บน disk จริง
                # flush() จะสั่งว่า: "เขียนเลยตอนนี้นะ!"
                temp_file.flush()
                # ย้าย “ตำแหน่ง pointer” ในไฟล์ — ว่าเราจะเริ่มอ่านหรือเขียนตรงไหน
                # หลังจากเขียนไฟล์เสร็จ pointer จะอยู่ "ท้ายไฟล์"
                # ถ้าจะเปิดอ่านต่อ (หรือใช้ file object เดิมอ่าน) ต้อง “ย้อนกลับไปที่ต้นไฟล์” → seek(0)
                temp_file.seek(0)
                wb = load_workbook(temp_file.name)
                ws = wb.active
                vstSettings = []
                header_row = [cell.value for cell in ws[2]]
                menu_names = header_row[3:]  # คอลัมน์เมนูเริ่มจาก index 4
                # ดึง SystemMenu ทั้งหมด
                system_menus = {menu.name: menu for menu in SystemMenu.objects(name__in=menu_names)}
                inserted_codes = set()
                rowCount = 3
                for row in ws.iter_rows(min_row=3):
                    code = str(row[0].value).strip() if row[0].value else None
                    if not code:
                        continue  # skip ถ้าไม่มี code 
                    if code in inserted_codes:
                        messages.error(request, f"Duplicate code: {code} in row ({rowCount})")# skip ถ้า code ซ้ำในไฟล์เดิม
                        return HttpResponseRedirect(reverse('indexSettingVst'))
                    user:User = User.objects.filter(code=code).first()
                    if not user:
                        messages.error(request, f"User not found in row ({rowCount}) with code: {code}")
                        return HttpResponseRedirect(reverse('indexSettingVst'))
                    checkVstSetting: VisitorSetting = VisitorSetting.objects.filter(user = user).first()
                    if checkVstSetting:
                        if checkVstSetting.isActive == True:
                            # -- ถ้ายัง active ให้ show error
                            messages.error(request, f"Duplicate code: {code} in row ({rowCount})")
                            return HttpResponseRedirect(reverse('indexSettingVst'))
                        else:
                            # -- ถ้ามีข้อมูลใน db แล้ว ไม่ active false ให้ลบข้อมูลออกจาก db แล้ว save ของใหม่จาก excel ไปเลย
                            checkVstSetting.delete()
                    selected_menu_names = []
                    for i, cell in enumerate(row[3:], start=3):
                        if str(cell.value).strip() == "1":
                            selected_menu_names.append(header_row[i])
                    if not selected_menu_names:
                        continue  # ข้ามถ้าไม่ได้เลือกเมนู
                    menu_refs = [system_menus[name] for name in selected_menu_names if name in system_menus]
                    # สร้าง Visitor
                    setting: VisitorSetting = VisitorSetting()
                    setting.user=user
                    setting.menus=menu_refs
                    setting.isActive=True
                    setting.isAdmin=False
                    setting.note=''
                    currentUser: User = request.currentUser
                    if currentUser:
                        uCreate = UserSnapshot().UserToSnapshot(currentUser)
                        if uCreate:
                            setting.createBy=uCreate
                    setting.createDate=timezone.now()
                    vstSettings.append(setting)
                    inserted_codes.add(code)
                    rowCount += 1
                # MongoEngine สามารถ insert หลาย document พร้อมกัน
                # ทำงานเร็วกว่า for setting in orgSettings: setting.save
                VisitorSetting.objects.insert(vstSettings)
            messages.success(request, 'Save Success')
            return HttpResponseRedirect(reverse('indexSettingVst'))
        except Exception as e:
            print(e)
            messages.error(request, str(e))
            return HttpResponseRedirect(reverse('importSettingOrg'))
    else: 
        return render(request, 'setting_visitor/importSetting.html')
    
@requiredLogin
def exportExcelTemplate(request: HttpRequest):
    try:
        visitorSetting: VisitorSetting = VisitorSetting.objects.filter(isActive = True)
        excel = CreateExcelTemplateSetting("Visitor Setting", "app_visitor", visitorSetting)
        if excel:
            # สร้าง response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="visitor-setting.xlsx"'
            # save workbook ลง response
            excel.save(response)
            return response
        messages.error(request, "excel template data not found")
        return HttpResponseRedirect(reverse('indexSettingVst'))
    except Exception as e:
        print(e)
        messages.error(request, str(e))
        return HttpResponseRedirect(reverse('indexSettingVst'))
    
def deleteVstSettingByUser(requester: User, user: User):
    vstSetting: VisitorSetting = VisitorSetting.objects.filter(user = user.id).first()
    if vstSetting:
        vstSetting.isActive = False
        vstSetting.updateDate = timezone.now()
        if requester:
            uUpdate = UserSnapshot().UserToSnapshot(requester)
            if uUpdate:
                vstSetting.updateBy = uUpdate
        vstSetting.save()